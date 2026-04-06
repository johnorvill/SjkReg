from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSM = Path(r"H:\SjkReg-FINAL-AoH-Nytt (test).xlsm")
OUTPUT_JSON = ROOT / "data" / "hundar.json"


@dataclass
class DogRow:
    sjukdom: str
    omrade: str
    regnr: str
    hundens_namn: str
    fodd: str | None
    kullnr: str
    fader_regnr: str
    faders_namn: str
    mor_regnr: str
    mor_namn: str


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def extract_rows() -> list[dict[str, Any]]:
    workbook = load_workbook(SOURCE_XLSM, data_only=True, read_only=True)
    sheet = workbook["Hundar"]
    rows: list[dict[str, Any]] = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = DogRow(
            sjukdom=stringify(values[0]),
            omrade=stringify(values[1]),
            regnr=stringify(values[2]).upper(),
            hundens_namn=stringify(values[3]),
            fodd=stringify(values[4]) or None,
            kullnr=stringify(values[5]).upper(),
            fader_regnr=stringify(values[6]).upper(),
            faders_namn=stringify(values[7]),
            mor_regnr=stringify(values[8]).upper(),
            mor_namn=stringify(values[9]),
        )
        rows.append(asdict(row))

    workbook.close()
    return rows


def main() -> None:
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    rows = extract_rows()
    payload = {
        "source": str(SOURCE_XLSM),
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "count": len(rows),
        "rows": rows,
    }
    OUTPUT_JSON.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {len(rows)} rows to {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
